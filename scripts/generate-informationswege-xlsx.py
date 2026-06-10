#!/usr/bin/env python3
"""
Generate public/cannabismythen-informationswege.xlsx — the Informationswege
(information-channels) download offered in the Daten-Explorer when the
dataset toggle is set to "Informationswege".

Mirrors the conventions of the hand-built myth workbook
(docs/cannabismythen.xlsx): a "Daten" sheet (long-format data lookup,
title row -> dark header row -> one row per channel x group) and an
"Erläuterungen" sheet (term -> explanation pairs).

Sources of truth (read-only):
  - public/data/information-sources.json     -> the data the live dashboard renders
  - src/content/dashboard-definitionen.json  -> on-site indicator/group definitions

The underlying values are derived from the CaRM Abschlussbericht, Tab. 4.12–4.15.

Re-run after the data or definitions change:

    pip install openpyxl
    python3 scripts/generate-informationswege-xlsx.py
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_DATA = os.path.join(ROOT, "public", "data", "information-sources.json")
SRC_DEFS = os.path.join(ROOT, "src", "content", "dashboard-definitionen.json")
OUT = os.path.join(ROOT, "public", "cannabismythen-informationswege.xlsx")

with open(SRC_DATA, encoding="utf-8") as fh:
    data = json.load(fh)
with open(SRC_DEFS, encoding="utf-8") as fh:
    defs = json.load(fh)

# Canonical group order. Labels come from the information-sources data itself
# (the same labels the live Informationswege view uses) so the Daten sheet,
# the Erläuterungen sheet, and the CSV export all read identically.
GROUP_ORDER = ["adults", "minors", "consumers", "young_adults", "parents"]
group_label = {g["id"]: g["name"] for g in data["sourceGroups"]}
cat_label = {c["id"]: c["name"] for c in data["sourceCategories"]}

# metric key -> Daten-sheet column header (matches the live indicator labels).
METRICS = ["search", "perception", "trust", "prevention"]
METRIC_HEADER = {
    "search": "Suche (%)",
    "perception": "Wahrnehmung (%)",
    "trust": "Vertrauen (Punkte)",
    "prevention": "Präventionspotential (Punkte)",
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF2C4A6B")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="FF2C4A6B")
TERM_FONT = Font(bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def metric_value(metric_key, gid, sid):
    """Return the value for source `sid` (int) in group `gid` for a metric,
    or None when that channel was not measured for this metric."""
    table = data["metrics"][metric_key]["data"].get(gid, {})
    if str(sid) in table:
        return table[str(sid)]
    if sid in table:
        return table[sid]
    return None


wb = Workbook()

# ---------------------------------------------------------------------------
# Sheet 1: Daten — long format, one row per channel x group.
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Daten"
ws["A1"] = (
    "Daten-Lookup: Informationskanal × Zielgruppe × Indikatoren "
    "(Quelle: information-sources.json / CaRM-Abschlussbericht, Tab. 4.12–4.15)"
)
ws["A1"].font = TITLE_FONT

headers = ["source_id", "Kanal", "Kategorie", "group_id", "Zielgruppe"] + [
    METRIC_HEADER[m] for m in METRICS
]
ws.append(headers)  # row 2
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=2, column=col)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

for src in data["sources"]:
    sid = src["id"]
    name = (src.get("name") or "").strip()  # children carry a leading "↳"
    cat = cat_label.get(src.get("category"), src.get("category", ""))
    for gid in GROUP_ORDER:
        row = [sid, name, cat, gid, group_label[gid]]
        for m in METRICS:
            row.append(metric_value(m, gid, sid))
        ws.append(row)

# One-decimal number format on the four value columns (F–I).
for r in range(3, ws.max_row + 1):
    for c in range(6, 6 + len(METRICS)):
        cell = ws.cell(row=r, column=c)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.0"

ws.freeze_panes = "A3"
for col, width in {
    "A": 9, "B": 34, "C": 22, "D": 13, "E": 24,
    "F": 11, "G": 16, "H": 16, "I": 24,
}.items():
    ws.column_dimensions[col].width = width

# ---------------------------------------------------------------------------
# Sheet 2: Erläuterungen — terms + definitions (reuses on-site definitions).
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Erläuterungen")
ws2["A1"] = "Erläuterungen — Begriffe und Lesehinweise zu den Informationswegen"
ws2["A1"].font = TITLE_FONT

rows = [
    (
        "Worum geht es?",
        "Diese Datei zeigt Ergebnisse des Forschungsprojekts CaRM "
        "(Cannabis – Risiken und Mythen, ISD Hamburg): die Wege, über die sich "
        "die cannabispräventiven Zielgruppen zu Gesundheitsthemen (u. a. Cannabis) "
        "informieren. Für jeden Informationskanal werden vier Kennwerte je "
        "Zielgruppe ausgewiesen.",
    ),
    (
        "Datengrundlage",
        "Bevölkerungsbefragung in Deutschland im Rahmen des Projekts CaRM "
        "(ISD Hamburg, 2025). Die Stichprobengrößen je Zielgruppe stehen unten; "
        "Volljährige und Minderjährige sind gewichtet (nach Geschlecht, Alter, "
        "Bildung und Migrationshintergrund).",
    ),
    ("Zielgruppen", "Die Kennwerte werden für fünf Zielgruppen ausgewiesen:"),
]

# Group definitions (label · n -> definition). Label from the data file,
# sample size + definition from the on-site definitions singleton.
for gid in GROUP_ORDER:
    g = defs["groups"][gid]
    term = group_label[gid]
    if g.get("sampleSize"):
        term = f"{group_label[gid]} · {g['sampleSize']}"
    rows.append((term, g["definition"]))

# Indicator definitions (Suche / Wahrnehmung / Vertrauen / Präventionspotential).
for key in METRICS:
    d = defs["sourcesIndicators"][key]
    definition = d["definition"]
    if d.get("scale"):
        definition = f"{definition} Skala: {d['scale']}."
    rows.append((d["label"], definition))

rows.extend([
    (
        "Unterkanäle & Mehrfachnennungen",
        "Einige Kanäle sind in Unterkanäle aufgegliedert (mit ↳ markiert, "
        "z. B. „persönlich“ oder „Internet“). Bei Suche und Wahrnehmung waren "
        "Mehrfachnennungen möglich. Leere Zellen bedeuten, dass der Kennwert für "
        "diesen Kanal nicht erhoben wurde.",
    ),
    (
        "Lesart der Stichproben",
        "Die Stichproben beziehen sich auf Erwachsene 18–70 und Minderjährige "
        "16–17 in einer Bevölkerungsbefragung in Deutschland — nicht auf „die "
        "Bevölkerung“ im Sinne der gesamten 80 Mio. Menschen. Bitte diese Lesart "
        "bei der Weiterverwendung beibehalten.",
    ),
    (
        "Quelle",
        "ISD Hamburg, Forschungsprojekt CaRM (Cannabis – Risiken und Mythen), "
        "Abschlussbericht, Tabellen 4.12–4.15. cannabismythen.de",
    ),
])

r = 3  # leave row 2 blank under the title
for term, explanation in rows:
    ws2.cell(row=r, column=1, value=term).font = TERM_FONT
    ws2.cell(row=r, column=1).alignment = WRAP_TOP
    c = ws2.cell(row=r, column=2, value=explanation)
    c.alignment = WRAP_TOP
    r += 1

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 90

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  Daten: {ws.max_row - 2} rows ({len(data['sources'])} channels × {len(GROUP_ORDER)} groups)")
print(f"  Erläuterungen: {len(rows)} entries")
